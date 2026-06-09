"""Build the AMG monthly investment report: formatted multi-tab Excel.

Tabs:
    1. Summary       — plan NAV, weighted return, allocation vs target by asset class
    2. By Manager    — full manager detail with drift and monthly return
    3. Reconciliation— custodian vs manager status with exceptions on top

Usage:
    python exports/build_monthly_report.py [--month 2026-06]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from recon.reconcile_performance import reconcile  # noqa: E402

DATA = ROOT / "data" / "portfolio_monthly.csv"

DARK = "0B5394"
HEADER_FILL = PatternFill("solid", fgColor=DARK)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=DARK)
WARN_FILL = PatternFill("solid", fgColor="FCE5CD")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(horizontal="center")


def autofit(ws) -> None:
    for col_cells in ws.columns:
        width = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width + 3


def write_frame(ws, frame: pd.DataFrame, start_row: int = 1,
                num_formats: dict[str, str] | None = None,
                warn_col: str | None = None, warn_values: set | None = None) -> None:
    num_formats = num_formats or {}
    for r_idx, row in enumerate(dataframe_to_rows(frame, index=False, header=True),
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            if r_idx > start_row:
                col_name = frame.columns[c_idx - 1]
                if col_name in num_formats:
                    cell.number_format = num_formats[col_name]
                if warn_col and frame.iloc[r_idx - start_row - 1][warn_col] in (warn_values or set()):
                    cell.fill = WARN_FILL
    style_header(ws, start_row, len(frame.columns))
    autofit(ws)


def build(month: str | None, out: Path) -> Path:
    df = pd.read_csv(DATA).drop_duplicates(subset=["manager", "month"], keep="first")
    month = month or sorted(df["month"].unique())[-1]
    cur = df[df["month"] == month].copy()
    if cur.empty:
        raise SystemExit(f"No data for {month}. Available: {sorted(df['month'].unique())[-6:]}")

    total_nav = cur["nav_usd_m"].sum()
    cur["actual_weight_pct"] = cur["nav_usd_m"] / total_nav * 100
    cur["drift_pct"] = cur["actual_weight_pct"] - cur["target_weight_pct"]
    wavg = (cur["custodian_return_pct"] * cur["nav_usd_m"]).sum() / total_nav

    wb = Workbook()

    # Tab 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"AMG Monthly Investment Report — {month}"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Plan NAV ($M)"; ws["B3"] = round(total_nav, 1)
    ws["A4"] = "Weighted monthly return (%)"; ws["B4"] = round(wavg, 2)
    ws["A5"] = "Managers reporting"; ws["B5"] = int(cur["manager"].nunique())
    for r in range(3, 6):
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER

    ac = (cur.groupby("asset_class")
             .agg(nav_usd_m=("nav_usd_m", "sum"),
                  target_pct=("target_weight_pct", "sum"))
             .reset_index())
    ac["actual_pct"] = ac["nav_usd_m"] / total_nav * 100
    ac["drift_pts"] = ac["actual_pct"] - ac["target_pct"]
    ws["A7"] = "Allocation vs target"; ws["A7"].font = Font(bold=True, color=DARK)
    write_frame(ws, ac.round(2), start_row=8,
                num_formats={"nav_usd_m": "#,##0", "target_pct": "0.0",
                             "actual_pct": "0.00", "drift_pts": "0.00"})

    # Tab 2 — By Manager
    ws2 = wb.create_sheet("By Manager")
    detail = cur[["manager", "asset_class", "sub_class", "nav_usd_m",
                  "target_weight_pct", "actual_weight_pct", "drift_pct",
                  "custodian_return_pct", "reporting_lag_months"]].round(3)
    write_frame(ws2, detail,
                num_formats={"nav_usd_m": "#,##0", "target_weight_pct": "0.0",
                             "actual_weight_pct": "0.00", "drift_pct": "0.00",
                             "custodian_return_pct": "0.00"})

    # Tab 3 — Reconciliation
    ws3 = wb.create_sheet("Reconciliation")
    recon = reconcile(pd.read_csv(DATA), month).round(3)
    write_frame(ws3, recon, warn_col="status",
                warn_values={"BREAK", "MISSING DATA", "MINOR"},
                num_formats={"custodian_return_pct": "0.00",
                             "manager_return_pct": "0.00", "diff_bps": "0.0"})

    wb.save(out)
    return out


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--month", default=None, help="e.g. 2026-06 (default: latest)")
    args = parser.parse_args()
    out = ROOT / "exports" / f"amg_monthly_report_{args.month or 'latest'}.xlsx"
    print(f"✓ Report written to {build(args.month, out)}")
